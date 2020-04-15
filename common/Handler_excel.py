''' 
author:紫夏
Time:2020/3/16 0:18
'''
import openpyxl
class Handler_Excel():

    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname

    def read_data(self):
        wb=openpyxl.load_workbook(self.filename)
        sh=wb[self.sheetname]
        rows_data=list(sh.rows)

        cases_data = []
        title=[]
        for i in rows_data[0]:
            title.append(i.value)


        for i in rows_data[1:]:
            values = []
            for j in i:
                values.append(j.value)
            case=dict(zip(title,values))
            cases_data.append(case)
        return cases_data

    def write_data(self,row,column,value):
        wb=openpyxl.load_workbook(self.filename)
        sh=wb[self.sheetname]
        sh.cell(row=row,column=column,value=value)
        wb.save(self.filename)



if __name__=='__main__':
    # excel1=Handler_Excel('cases.xlsx','register')
    # cases=excel1.read_data()
    # print(cases)

    excel=Handler_Excel('F:\python37test\py27_api_test\data\cases.xlsx','register')
    cases=excel.read_data()
    print(cases)